
from openpyxl import Workbook
from openpyxl.drawing.image import Image

from flask import current_app, request, send_from_directory, send_file
from flask_restx import Namespace, Resource

from library.utilities.inference import get_prediction
from PIL import Image as PILimg
from PIL import ImageOps
import os

import hashlib

################ Blueprint/Namespace Configuration ################
utils_api = Namespace('utilities_api', description='Utilities related operations')
json_parser = utils_api.parser()
json_parser.add_argument('results', location='json', type=list, required=True, help='Results JSON')

def resize_img(path):
    img = PILimg.open(path)
    img = ImageOps.fit(img, (100, 100), method=PILimg.Resampling.LANCZOS)
    img.save(path)

################ Global Variables ################
storage_path = current_app.config['STORAGE_FOLDER']
img_path = current_app.config['UPLOAD_FOLDER']
is_prod = os.getenv('FLASK_ENV') == 'production'
################ API Endpoints ################

@utils_api.route('/create_xlsx')
class XLSX_export(Resource):
    
    @utils_api.response(200, 'Success')
    @utils_api.response(400, 'Bad Request (Likely Invalid JSON)')
    @utils_api.response(500, 'Internal Server Error')
    @utils_api.expect(json_parser)
    @utils_api.doc(description="Creates an XLSX file from JSON prediction data")
    def post(self):
        try:
            if request.is_json:
                json_data = request.get_json()

                # Validate JSON
                for obj in json_data:
                    if "name" not in obj or "pred" not in obj:
                        return "Invalid JSON", 400
                
                # File name is based on the hash of the JSON data
                str_data = str(json_data)
                hash_object = hashlib.md5(str_data.encode())
                hash_hex = hash_object.hexdigest()

                # Create the file if it doesn't exist
                if not os.path.isfile(storage_path+hash_hex+".xlsx"):
                    wb = Workbook()
                    ws = wb.active
                    
                    # Write the data to the file
                    ws.append(["Filename","Image", "Predictions"])
                    for result in json_data:
                        filename = result["name"]
                        pred = [item for subpred in result["pred"] for item in subpred]
                        ws.append([filename]+[""]+pred)
                        if os.path.isfile(img_path+filename):
                            resize_img(img_path+filename)

                            img = Image(img_path+filename)
                            ws.add_image(img, "B"+str(ws.max_row))
                    
                    #Auto adjust column width A (FILENAME)
                    max_length = 0
                    for cell in ws['A']:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions['A'].width = adjusted_width

                    # Adjust column width B (IMAGE)
                    ws.column_dimensions['B'].width = 14 # Found via trial and error
                    for row in ws.iter_rows(min_row=2):
                        ws.row_dimensions[row[0].row].height = 80 # Found via trial and error

                    wb.save(storage_path+hash_hex+".xlsx")
                
                path = os.path.join(current_app.static_folder, 'storage/') # Yes I know this is inconsistent, This is the only way i could get this to work
                return send_file(os.path.join(path, hash_hex+".xlsx"), as_attachment=True)
                # return send_from_directory(path, hash_hex+".xlsx")
            else:
                return "Invalid JSON", 400
        except Exception as e:
            if is_prod:
                return "Internal Server Error", 500
            else:
                return str(e), 500
        
@utils_api.route('/create_csv')
class CSV_export(Resource):

    @utils_api.response(200, 'Success')
    @utils_api.response(400, 'Bad Request (Likely Invalid JSON)')
    @utils_api.response(500, 'Internal Server Error')
    @utils_api.expect(json_parser)
    @utils_api.doc(description="Creates a CSV file from JSON prediction data")
    def post(self):
        try:
            if request.is_json:
                json_data = request.get_json()

                for obj in json_data:
                    if "name" not in obj or "pred" not in obj:
                        return "Invalid JSON", 400
                str_data = str(json_data)
                hash_object = hashlib.md5(str_data.encode())
                hash_hex = hash_object.hexdigest()

                if not os.path.isfile(storage_path+hash_hex+".csv"):
                    with open(storage_path+hash_hex+".csv", 'w') as f:
                        f.write("Filename,Predictions,Label\n")
                        for result in json_data:
                            filename = result["name"]
                            pred = [item for subpred in result["pred"] for item in subpred]
                            # f.write(filename)
                            for i in range(0, len(pred), 2):
                                f.write(filename+","+pred[i]+","+pred[i+1])
                                f.write("\n")
                            # f.write("\n")
                            # f.write(filename+",".join(pred)+"\n")
                path = os.path.join(current_app.static_folder, 'storage/')
                return send_file(os.path.join(path, hash_hex+".csv"), as_attachment=True)
            else:
                return "Invalid JSON",400   
        except Exception as e:
            if is_prod:
                return "Internal Server Error", 500
            else:
                return str(e), 500
           

