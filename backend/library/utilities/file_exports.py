from fastapi import APIRouter, Request, Body
from fastapi.responses import FileResponse, ORJSONResponse, JSONResponse

from PIL import Image as PILimg
from PIL import ImageOps
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image

import hashlib
from ..config import Settings

################ Blueprint/Namespace Configuration ################

utils_api = APIRouter(tags=["Utilities"])

################ Global Variables ################

storage_path = Settings().STORAGE_FOLDER
img_path = Settings().UPLOAD_FOLDER
is_prod = Settings().FLASK_ENV == 'production'


################ Helper Functions ################

def resize_img(path):
    img = PILimg.open(path)
    img = ImageOps.fit(img, (100, 100), method=PILimg.Resampling.LANCZOS)
    img.save(path)

################ API Endpoints ################

@utils_api.post('/api/v1/create_xlsx', responses={200: {"description": "Success"}, 400: {"description": "Bad Request (Likely Invalid JSON)"}, 405: {"description": "Method Not Allowed"}, 500: {"description": "Internal Server Error"}}, tags=["Utilities"])
async def json_to_xlsx(request: Request, results: list[dict]= Body(...)):
    """
        Creates an XLSX file from JSON prediction data
    """
    
    try:
        # Validate JSON
        for obj in results:
            if "name" not in obj or "pred" not in obj:
                return ORJSONResponse(content={"error": "Invalid JSON"}, status_code=400)            
        
        # File name is based on the hash of the JSON data
        file_name = hashlib.md5(str(results).encode()).hexdigest()
        file_path = os.path.join(storage_path, file_name+".xlsx")

        # Create the file if it doesn't exist
        if not os.path.isfile(file_path):
            wb = Workbook()
            ws = wb.active
            
            # Write the data to the file
            ws.append(["Filename","Image", "Predictions"])
            for result in results:
                filename = result["name"]
                pred = [item for subpred in result["pred"] for item in subpred]
                ws.append([filename]+[""]+pred)
                if os.path.isfile(img_path+filename):
                    resize_img(img_path+filename)

                    img = Image(img_path+filename)
                    ws.add_image(img, "B"+str(ws.max_row))
            
            #Auto adjust column width A (FILENAME)
            max_length = 0
            for cell in ws['A']:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions['A'].width = adjusted_width

            # Adjust column width B (IMAGE)
            ws.column_dimensions['B'].width = 14 # Found via trial and error
            for row in ws.iter_rows(min_row=2):
                ws.row_dimensions[row[0].row].height = 80 # Found via trial and error

            wb.save(file_path)
        return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=file_name+".xlsx")
    except Exception as e:
        if is_prod:
            return ORJSONResponse(content={"error": "Internal Server Error"}, status_code=500)
        else:
            return JSONResponse(content={"error": str(e)}, status_code=500)

@utils_api.post('/api/v1/create_csv', responses={200: {"description": "Success"}, 400: {"description": "Bad Request (Likely Invalid JSON)"}, 405: {"description": "Method Not Allowed"}, 500: {"description": "Internal Server Error"}}, tags=["Utilities"])
async def json_to_csv(request: Request, results: list[dict]= Body(...)):
    """
        Creates a CSV file from JSON prediction data
    """
    try:
        for obj in results:
            if "name" not in obj or "pred" not in obj:
                return ORJSONResponse(content={"error": "Invalid JSON"}, status_code=400)    
            
        file_name = hashlib.md5(str(results).encode()).hexdigest()
        file_path = os.path.join(storage_path, file_name+".csv")

        if not os.path.isfile(file_path):
            with open(file_path, 'w') as f:
                f.write("Filename,Predictions,Label\n")
                for result in results:
                    filename = result["name"]
                    pred = [item for subpred in result["pred"] for item in subpred]
                    # f.write(filename)
                    for i in range(0, len(pred), 2):
                        f.write(filename+","+pred[i]+","+pred[i+1])
                        f.write("\n")
                    # f.write("\n")
                    # f.write(filename+",".join(pred)+"\n")
        return FileResponse(file_path, media_type="text/csv", filename=file_name+".csv")
    except Exception as e:
        if is_prod:
            return ORJSONResponse(content={"error": "Internal Server Error"}, status_code=500)
        else:
            return JSONResponse(content={"error": str(e)}, status_code=500)
           

